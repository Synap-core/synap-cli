"""
Build Review Builder
====================

Scaffolds the 4-sheet Excel deliverable for an SCF Build phase review.

Same structure as abstract-review-builder.py with Build-phase fields (Q1-Q17).
Customize FIELDS, CHECKLIST, COMPARABLES, INTEGRATION_AUDIT, BANNER_TEXT,
PROJECT_NAME, VERSION and run: python build-review-builder.py.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PROJECT_NAME = "ProjectX"
VERSION = "1"
OUT = f"/sessions/eloquent-inspiring-cannon/mnt/outputs/{PROJECT_NAME}_Build_Review_v{VERSION}.xlsx"

BANNER_TEXT = (
    f"{PROJECT_NAME} Build Review v{VERSION}: estimated XX to YY percent probability of acceptance "
    f"as currently drafted, lifting to AA to BB percent after the N fixes below. "
    f"Recommended budget anchor: $XXX,000 Integration Track, anchored on [Named Past Winner] ($YYK SCF #NN)."
)

# -------------------------------------------------------------------
# Sheet 1: Field-by-Field Review (Build phase Q1-Q17)
# -------------------------------------------------------------------
FIELDS = [
    ("Q1 Project Title", "STRONG", "Coherent avec Abstract.", "[Project Name]"),
    ("Q2 Submission Title (40 chars)", "SOLID", "Specifique.", "[Stellar-Native [Specific Build]]"),
    ("Q3 One Sentence (130 chars)", "WEAK", "Generic.", "[Verb] [defined offering] to help [audience] [problem] with [secret sauce]."),
    ("Q4 Project URL", "STRONG", "Valide.", "https://yourdomain.com/"),
    ("Q5 Code URL", "BLOCKER", "Empty Q5 is pre-screen rejection.", "https://github.com/[org]/[repo] with README + MIT + scaffold."),
    ("Q6 Video URL", "BLOCKER", "Empty Q6 is pre-screen rejection.", "https://youtube.com/watch?v=[id] with 5-segment script under 3 min."),
    ("Q7 Soroban", "SOLID", "Yes correct.", "Yes"),
    ("Q8 Product and Services", "WEAK", "Contracts non nommes.", "Component 1: [Contract.rs]. Stellar use: [item]. Impact: [...]"),
    ("Q9 Traction Evidence", "UNDERSELLING", "Manque LOIs nominees.", "Live URLs, metrics, named clients, recognitions."),
    ("Q10 Technical Architecture", "BLOCKER", "Empty.", "Public Notion or GitHub markdown with C4 L1+L2 + data flows."),
    ("Q11 Tranche intro", "WEAK", "Pas d'anchor.", "$X, three tranches of $X/3. NOT FUNDED: marketing. Anchored on [Named Past Winner]."),
    ("Q12 Tranche 1 MVP", "WEAK", "Vague.", "Deliverable: [contract]. Measure: testnet deployed at [addr]. Date: T+N weeks. Budget: $X."),
    ("Q13 Tranche 2 Testnet", "WEAK", "Vague.", "Same format per deliverable."),
    ("Q14 Tranche 3 Mainnet", "WEAK", "Vague.", "Same format + quantified mainnet targets."),
    ("Q15 Budget Total", "WEAK", "Drop to anchor.", "$[anchored amount]"),
    ("Q16 Go-To-Market", "UNDERSELLING", "Manque grant note.", "Phase 1: ... Grant note: SCF funds dev only, marketing self-funded."),
    ("Q17 Success Criteria", "WEAK", "Pas de quantification.", "MVP / Testnet / Mainnet + quantified targets + ecosystem + business + long-term."),
]

# Re-use the rest of the structure from abstract-review-builder.py (CHECKLIST,
# COMPARABLES, INTEGRATION_AUDIT, fills, build function). For brevity here
# we focus on the structural differences.

CHECKLIST = [
    ("BLOCKER", "Ship testnet contract + add URL to Q5", "Q5", "5-7 days", "Empty Q5 is pre-screen rejection.", ""),
    ("BLOCKER", "Record 3-min video + upload to Q6", "Q6", "6-8 hours", "Empty Q6 is pre-screen rejection.", ""),
    ("BLOCKER", "Publish Technical Architecture on Notion + add URL to Q10", "Q10", "4-6 hours", "Empty Q10 is pre-screen rejection.", ""),
    ("CRITICAL", "Add named LOIs/clients to Q9", "Q9", "1 day", "Named traction removes 'will this find users?' objection.", ""),
    ("CRITICAL", "Add grant note to Q16 ('SCF funds dev only')", "Q16", "15 min", "Blocks marketing-budget-misuse pattern.", ""),
    ("IMPORTANT", "Quantify Tranche 3 mainnet targets", "Q14", "30 min", "Reviewers need something to verify at final tranche.", ""),
    ("STRATEGIC", "Add 2-3 SCF Integration List items to Q8", "Q8", "30 min", "Strongest dossiers leverage 3-7 items.", ""),
]

COMPARABLES = [
    (f"{PROJECT_NAME} (current)", "SCF #44", "TBD", "TBD", "Application", "Profile.", "N/A current dossier."),
    ("The Signal", "SCF #42", "$121,000", "Won", "B2B marketplace", "2 doxxed cofounders, DealEscrow.rs, Atomic Splits.", "Anchor for $115-125K."),
    ("Bando", "SCF #42", "$75,000", "Won", "RWA Mexico", "Doxxed, brokerage partner, pilot.", "Anchor for $75-90K narrow pilots."),
    ("TheXBank", "SCF #40", "$89,500", "Won", "Africa banking", "Doxxed, country pilot.", "Anchor for emerging markets banking."),
    ("Latch + KMP", "SCF #41 RFP", "RFP", "Won", "Compliance tooling", "Compliance for Soroban ecosystem.", "Anchor for compliance dossiers."),
    ("Orion + OctoPos", "SCF #41 RFP", "RFP", "Won", "DeFi Positions API", "Open data API across DeFi.", "Anchor for indexer dossiers."),
    ("Sorotrack", "SCF #40+41+42", "$150K + $50K + $50K", "PRF 3x", "Tracker", "Solo dev by handle.", "Doxxing matters."),
    ("Reactor Trade", "SCF #41", "$115,000", "Rejected", "Trading", "Lost to Helix Labs same round.", "Audit competing teams."),
    ("Lend.xyz", "SCF #42", "$120,000", "Rejected", "RWA real estate", "Lost to Bando.", "Sharper sub-niche."),
    ("Tessera Labs", "SCF #42", "$150,000", "Rejected", "Confidential payments", "Anonymous + brand collision + duplicates X-Ray.", "Three patterns stacked."),
    ("PathPulse AI", "SCF #41/#42", "$137K / $128K", "Rejected 2x", "Computer-vision data", "Bolt-on without Stellar reason.", "Stellar-specific story needed."),
    ("Goated App", "SCF #41", "$120,000", "Rejected", "Predictions", "No licensing plan.", "Regulated category needs framework."),
    ("Kyros OS", "SCF #41", "$135,000", "Rejected", "Spanish tax", "No AEAT/Hacienda partner.", "Regulated needs partner."),
    ("Rebond (target)", "SCF #44", "$132,000", "TBD", "Green bonds", "4 LOIs, Art L.411-2.", "Anchor for $130K+ with LOIs."),
    ("For Yield (target)", "SCF #44", "$144,000", "TBD", "Regulated DeFi", "PSCA AMF, 7-8M EUR AUM.", "Anchor for $140K+ regulated."),
]

INTEGRATION_AUDIT = [
    ("Wallets", "Stellar Wallets Kit", "TO ADD", "Default multi-wallet connector.", "<1 day"),
    ("Wallets", "Privy", "OPTIONAL", "For non-crypto-native users.", "<1 day"),
    ("Wallets", "DFNS", "OPTIONAL", "For institutional MPC custody.", "1-5 days"),
    ("On/Off-ramping", "Anchor Platform", "TO ADD / NOT RELEVANT", "Add if fiat ramp needed.", "1+ month"),
    ("On/Off-ramping", "Bridge", "OPTIONAL", "Alternative ramp.", "1-5 days"),
    ("DeFi Yield", "DeFindex", "OPTIONAL", "For yield routing.", "1-2 weeks"),
    ("DeFi", "Blend v2", "OPTIONAL", "Lending pools.", "1+ month"),
    ("DeFi", "Aquarius", "OPTIONAL", "DEX/LP.", "<1 day"),
    ("DeFi", "Soroswap", "OPTIONAL", "DEX aggregation.", "1-2 weeks"),
    ("Cross-chain", "Allbridge", "OPTIONAL", "Stablecoin bridging.", "TBD"),
    ("Cross-chain", "Near Intents", "OPTIONAL", "Cross-chain execution.", "<1 day"),
    ("Payments", "Stellar Disbursement Platform", "OPTIONAL / NOT RELEVANT", "For bulk payouts.", "1-2 weeks"),
]

# Same styling and build() as abstract-review-builder.py
STATUS_FILLS = {
    "STRONG": PatternFill(start_color="C6EFCE", fill_type="solid"),
    "SOLID": PatternFill(start_color="E2F0D9", fill_type="solid"),
    "UNDERSELLING": PatternFill(start_color="FFEB9C", fill_type="solid"),
    "WEAK": PatternFill(start_color="FCD5B5", fill_type="solid"),
    "BLOCKER": PatternFill(start_color="FFC7CE", fill_type="solid"),
}
PRIORITY_FILLS = {
    "BLOCKER": PatternFill(start_color="FFC7CE", fill_type="solid"),
    "CRITICAL": PatternFill(start_color="FCD5B5", fill_type="solid"),
    "IMPORTANT": PatternFill(start_color="FFEB9C", fill_type="solid"),
    "STRATEGIC": PatternFill(start_color="DDEBF7", fill_type="solid"),
}
VERDICT_FILLS = {
    "Won": PatternFill(start_color="C6EFCE", fill_type="solid"),
    "Rejected": PatternFill(start_color="FCD5B5", fill_type="solid"),
    "PRF 3x": PatternFill(start_color="FFC7CE", fill_type="solid"),
    "Rejected 2x": PatternFill(start_color="FCD5B5", fill_type="solid"),
    "TBD": PatternFill(start_color="FFFFFF", fill_type="solid"),
}
STATUS_AUDIT_FILLS = {
    "MENTIONED": PatternFill(start_color="C6EFCE", fill_type="solid"),
    "TO ADD": PatternFill(start_color="FFEB9C", fill_type="solid"),
    "OPTIONAL": PatternFill(start_color="DDEBF7", fill_type="solid"),
    "NOT RELEVANT": PatternFill(start_color="F2F2F2", fill_type="solid"),
    "TO ADD / NOT RELEVANT": PatternFill(start_color="FFEB9C", fill_type="solid"),
    "OPTIONAL / NOT RELEVANT": PatternFill(start_color="DDEBF7", fill_type="solid"),
}
BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="0E1116", fill_type="solid")


def add_banner(ws, text, n_cols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    c = ws.cell(row=1, column=1, value=text)
    c.font = Font(bold=True, color="FFFFFF", size=12)
    c.fill = HEADER_FILL
    c.alignment = Alignment(wrap_text=True, vertical='center')
    ws.row_dimensions[1].height = 80


def add_header_row(ws, headers, row=2):
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical='center')
        cell.border = BORDER


def write_data_row(ws, row_idx, values, fills=None):
    for col, val in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col, value=val)
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        cell.border = BORDER
        if fills and col in fills:
            cell.fill = fills[col]


def set_widths(ws, widths):
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w


def build():
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Build Review"
    add_banner(ws1, BANNER_TEXT, 4)
    add_header_row(ws1, ["Field", "Status", "Review", "Rewrite / Action"], row=2)
    set_widths(ws1, [26, 20, 70, 80])
    for i, (field, status, review, rewrite) in enumerate(FIELDS, start=3):
        write_data_row(ws1, i, [field, status, review, rewrite],
                       fills={2: STATUS_FILLS.get(status, PatternFill())})

    ws2 = wb.create_sheet("Checklist Pre-Submit")
    add_banner(ws2, f"{PROJECT_NAME} Pre-Submission Checklist v{VERSION}", 6)
    add_header_row(ws2, ["Priority", "Item", "Field", "Effort", "Why it matters", "Status"], row=2)
    set_widths(ws2, [14, 50, 14, 12, 60, 14])
    for i, (prio, item, field, effort, why, status) in enumerate(CHECKLIST, start=3):
        write_data_row(ws2, i, [prio, item, field, effort, why, status],
                       fills={1: PRIORITY_FILLS.get(prio, PatternFill())})

    ws3 = wb.create_sheet("Comparables Corpus")
    add_banner(ws3, f"{PROJECT_NAME} Corpus Comparables", 7)
    add_header_row(ws3, ["Project", "Round", "Budget", "Verdict", "Category", "Profile", "Lesson"], row=2)
    set_widths(ws3, [26, 14, 18, 18, 26, 40, 50])
    for i, (proj, rnd, budget, verdict, cat, profile, lesson) in enumerate(COMPARABLES, start=3):
        write_data_row(ws3, i, [proj, rnd, budget, verdict, cat, profile, lesson],
                       fills={4: VERDICT_FILLS.get(verdict, PatternFill())})

    ws4 = wb.create_sheet("Audit Integration List")
    add_banner(ws4, f"{PROJECT_NAME} SCF Integration List Audit", 5)
    add_header_row(ws4, ["Category", "Item", "Status in dossier", "Recommendation", "Effort"], row=2)
    set_widths(ws4, [20, 30, 24, 60, 14])
    for i, (cat, item, status, reco, effort) in enumerate(INTEGRATION_AUDIT, start=3):
        write_data_row(ws4, i, [cat, item, status, reco, effort],
                       fills={3: STATUS_AUDIT_FILLS.get(status, PatternFill())})

    wb.save(OUT)
    print(f"Saved: {OUT}")
    from openpyxl import load_workbook
    wb2 = load_workbook(OUT)
    em = sum(1 for ws in wb2.worksheets for row in ws.iter_rows() for cell in row
             if cell.value and isinstance(cell.value, str) and ", " in cell.value)
    print(f"em-dash cells: {em} (expect 0)")
    print(f"Sheets: {wb2.sheetnames}")


if __name__ == "__main__":
    build()
