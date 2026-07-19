"""
Abstract Review Builder
=======================

Scaffolds the 4-sheet Excel deliverable for an SCF Abstract review.

Usage:
    Customize the FIELDS list, CHECKLIST, COMPARABLES, INTEGRATION_AUDIT,
    BANNER_TEXT and PROJECT_NAME for the dossier being reviewed,
    then run: python abstract-review-builder.py

Outputs:
    {PROJECT_NAME}_Abstract_Review_v{VERSION}.xlsx

Conventions:
    - Column D (rewrites) MUST be paste-ready English.
    - Column C (commentary) can be French or English depending on client.
    - NEVER use "kill signature" anywhere in any cell.
    - NEVER use em-dashes anywhere in any cell.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys

PROJECT_NAME = "ProjectX"
VERSION = "1"
OUT = f"/sessions/eloquent-inspiring-cannon/mnt/outputs/{PROJECT_NAME}_Abstract_Review_v{VERSION}.xlsx"

BANNER_TEXT = (
    f"{PROJECT_NAME} Abstract Review v{VERSION}: estimated XX to YY percent probability of acceptance "
    f"as currently drafted, lifting to AA to BB percent after the N fixes below. "
    f"Recommended budget anchor: $XXX,000 Integration Track, anchored on [Named Past Winner] ($YYK SCF #NN)."
)

# -------------------------------------------------------------------
# Sheet 1: Field-by-Field Review
# -------------------------------------------------------------------
# Format: (Field, Status, Review (FR or EN), Rewrite/Action (EN paste-ready))
FIELDS = [
    ("Q1 Project Title", "STRONG", "Le titre est specifique...", "[Project Name]"),
    ("Q2 Description", "WEAK", "Le Q2 manque...", "[Project Name] is [the positioning], serving [segment]. The product: ... Why Stellar: ... Traction: ... Day-1 distribution: ..."),
    ("Q3 Project Category", "SOLID", "Bien choisi...", "Application"),
    ("Q4 Traction", "UNDERSELLING", "Le dossier ne mentionne pas...", "PRODUCT: [url, x, github, metrics]. PIPELINE: ... PARTNERS: ... RECOGNITIONS: ..."),
    ("Q5 Website URL", "STRONG", "URL valide", "https://yourdomain.com/"),
    ("Q6 Integration Description", "WEAK", "Aucun nom de contrat...", "[Project] uses Soroban as ... Phase 1 ... Phase 2 ... Phase 3 ... Why Stellar ... Integration List items: ..."),
    ("Q7 Track", "SOLID", "Integration Track est defendable", "Integration Track"),
    ("Q8 Thumbnail", "BLOCKER", "Q8 vide. Five past dossiers died in Information Collection status...", "Upload 1200x630 thumbnail. Source: existing og-image.webp on website."),
    ("Q9 Submitter type", "SOLID", "Entity est correct", "Entity"),
    ("Q10 Team", "UNDERSELLING", "Le Q10 sous-vend la team. LinkedIn credentials a integrer...", "[Founder 1 with LinkedIn URL + role + experience + recognition...]"),
    ("Q11 Number of team members", "SOLID", "Coherent avec Q10", "3"),
    ("Q12 Discord usernames", "WEAK", "Separateur slash au lieu de virgule par instruction SCF", "username1, username2, username3"),
]

# -------------------------------------------------------------------
# Sheet 2: Checklist Pre-Submit
# -------------------------------------------------------------------
# Format: (Priority, Item, Field, Effort, Why it matters, Status)
CHECKLIST = [
    ("BLOCKER", "Upload thumbnail Q8 from existing og-image.webp", "Q8", "15 min", "Empty Q8 risks Information Collection status. Five past dossiers died this way.", ""),
    ("CRITICAL", "Add LinkedIn URLs for all three co-founders in Q10", "Q10", "30 min", "Anonymous founders is the most reliable predictor of rejection in our database.", ""),
    ("CRITICAL", "Enrich Q10 with named credentials (schools, prior roles, recognitions)", "Q10", "1h", "Surface signals the team underestimates: IMC Top 0.1 pct, Bitget COO Apprentice, Bloomberg Market Concepts.", ""),
    ("IMPORTANT", "Rewrite Q2 to surface Stellar positioning", "Q2", "1h", "No mention of Stellar in Q2 is a pattern shared by several rejected dossiers.", ""),
    ("IMPORTANT", "Drop budget from $150K to anchor on Bando + The Signal", "Q15 in Build phase", "5 min", "Max ask without scope triggers heightened scrutiny.", ""),
    ("STRATEGIC", "Add 2-3 SCF Integration List items to Q6", "Q6", "30 min", "Strongest dossiers leverage 3-7 Integration List items.", ""),
    ("STRATEGIC", "Fix Q12 Discord separator from slash to comma", "Q12", "5 min", "Small detail signals attention to instructions.", ""),
]

# -------------------------------------------------------------------
# Sheet 3: Comparables Corpus
# -------------------------------------------------------------------
# Format: (Project, Round, Budget, Verdict, Category, Profile, Lesson)
COMPARABLES = [
    (f"{PROJECT_NAME} (current dossier)", "SCF #44", "TBD", "TBD", "Application / Financial Protocol", "Profile snapshot one-liner.", "N/A this is the current dossier."),
    ("The Signal", "SCF #42", "$121,000", "Won", "B2B marketplace + escrow", "2 doxxed cofounders, DealEscrow.rs, Atomic Splits, B2B verticals named.", "Closest budget anchor for $115K-$125K dossiers."),
    ("Bando", "SCF #42", "$75,000", "Won", "RWA real estate Mexico", "Doxxed team, named brokerage partner, country pilot.", "Anchor for narrow-pilot dossiers at $75K-$90K."),
    ("TheXBank", "SCF #40", "$89,500", "Won", "Africa banking", "Doxxed team, country pilot named.", "Anchor for emerging-market banking $80K-$100K."),
    ("AION_FI Protocol", "SCF #40", "$146,500", "Panel Review Failed", "Credit cards", "Anonymous team, no BIN sponsor.", "Same risk if regulatory partners not named."),
    ("Stablpay", "SCF #40", "$129,100", "Panel Review Failed", "Stablecoin rail India", "No FIU partner named.", "Same risk for regulated geographies without partners."),
    ("Sorotrack", "SCF #40", "$150,000", "Panel Review Failed (3x)", "Stellar tracker", "Solo dev only by handle 'Gemy'.", "Doxxing depth matters at high budget."),
    ("Stableyard", "SCF #40", "$119,500", "Rejected", "Asia QR stablecoin", "Brand collision + vague TAM.", "Brand search before submitting."),
    ("Lend.xyz", "SCF #42", "$120,000", "Rejected", "RWA real estate", "Lost to Bando same round.", "Audit prior winners before submitting same vertical."),
    ("Tessera Labs", "SCF #42", "$150,000", "Rejected", "Confidential payments", "Anonymous + brand collision + duplicates X-Ray.", "Three patterns in one dossier."),
    ("$NRG Token", "SCF #41", "$150,000", "Rejected", "Fan token rewards", "Earn-to-X + bolt-on + max ask.", "Pattern stack."),
    ("StellarRead", "SCF #40", "$150,000", "Panel Review Failed", "Read-to-earn", "Earn-to-X at max ask.", "Earn-to-X has 100 pct rejection rate."),
    ("Easner", "SCF #40", "$150,000", "Panel Review Failed", "Stripe for Stablecoins", "No banking partner.", "Generic positioning + no partner."),
    ("Rebond (target)", "SCF #44", "$132,000", "TBD", "Green bond RWA", "Doxxed team, 4 LOIs, Art L.411-2 framework.", "Anchor for dossiers with named LOIs at $130K."),
    ("For Yield (target)", "SCF #44", "$144,000", "TBD", "Regulated DeFi yield", "11-team, PSCA AMF filed, 7-8M EUR AUM.", "Anchor for regulated DeFi at $140-145K."),
]

# -------------------------------------------------------------------
# Sheet 4: Integration List Audit
# -------------------------------------------------------------------
# Format: (Category, Item, Status, Recommendation, Effort)
INTEGRATION_AUDIT = [
    ("On/Off-ramping", "Anchor Platform", "TO ADD", "Mandatory for EUR-to-USDC fiat ramp.", "1+ month"),
    ("On/Off-ramping", "Bridge", "OPTIONAL", "For Stripe-style multi-currency.", "1-5 days"),
    ("On/Off-ramping", "MoneyGram", "NOT RELEVANT", "Targets emerging-market cash, not aligned.", "1+ month"),
    ("On/Off-ramping", "Mercuryo", "OPTIONAL", "Lower-effort fiat-crypto ramp alternative.", "1-2 weeks"),
    ("DeFi Yield", "DeFindex", "MENTIONED / TO ADD", "For yield routing if scope includes yield product.", "1-2 weeks"),
    ("DeFi Protocol", "Blend v2", "OPTIONAL", "Lending pool primitive.", "1+ month"),
    ("DeFi Protocol", "Aquarius", "OPTIONAL", "DEX and LP pools.", "<1 day"),
    ("DeFi Protocol", "Soroswap", "OPTIONAL", "DEX aggregation.", "1-2 weeks"),
    ("DeFi Protocol", "Stellar Broker", "OPTIONAL", "Alternative router.", "1-5 days"),
    ("Cross-chain", "Allbridge", "OPTIONAL", "EVM stablecoin bridging.", "TBD"),
    ("Cross-chain", "Near Intents", "OPTIONAL", "Cross-chain execution layer.", "<1 day"),
    ("Cross-chain", "Axelar", "OPTIONAL", "General messaging.", "TBD"),
    ("Payments", "Stellar Disbursement Platform", "NOT RELEVANT / OPTIONAL", "Use for bulk payouts.", "1-2 weeks"),
    ("Wallets", "Stellar Wallets Kit", "TO ADD", "Default multi-wallet connector.", "<1 day"),
    ("Wallets", "Freighter Connect", "OPTIONAL", "SDF-maintained browser wallet.", "1-2 weeks"),
    ("Wallets", "Privy", "TO ADD", "Mandatory for non-crypto-native distribution (CGPs, family offices).", "<1 day"),
    ("Wallets", "DFNS", "OPTIONAL", "Institutional MPC custody.", "1-5 days"),
]


# ===================================================================
# Excel generation
# ===================================================================

STATUS_FILLS = {
    "STRONG": PatternFill(start_color="C6EFCE", fill_type="solid"),
    "SOLID": PatternFill(start_color="E2F0D9", fill_type="solid"),
    "UNDERSELLING": PatternFill(start_color="FFEB9C", fill_type="solid"),
    "WEAK": PatternFill(start_color="FCD5B5", fill_type="solid"),
    "BLOCKER": PatternFill(start_color="FFC7CE", fill_type="solid"),
}

PRIORITY_FILLS = {
    "BLOCKER": PatternFill(start_color="FFC7CE", fill_type="solid"),
    "CRITICAL": PatternFill(start_color="FCD5B5", fill_type="solid"),
    "IMPORTANT": PatternFill(start_color="FFEB9C", fill_type="solid"),
    "STRATEGIC": PatternFill(start_color="DDEBF7", fill_type="solid"),
}

VERDICT_FILLS = {
    "Won": PatternFill(start_color="C6EFCE", fill_type="solid"),
    "Rejected": PatternFill(start_color="FCD5B5", fill_type="solid"),
    "Panel Review Failed": PatternFill(start_color="FFC7CE", fill_type="solid"),
    "Panel Review Failed (3x)": PatternFill(start_color="FFC7CE", fill_type="solid"),
    "Information Collection": PatternFill(start_color="F2F2F2", fill_type="solid"),
    "TBD": PatternFill(start_color="FFFFFF", fill_type="solid"),
}

STATUS_AUDIT_FILLS = {
    "MENTIONED": PatternFill(start_color="C6EFCE", fill_type="solid"),
    "TO ADD": PatternFill(start_color="FFEB9C", fill_type="solid"),
    "OPTIONAL": PatternFill(start_color="DDEBF7", fill_type="solid"),
    "NOT RELEVANT": PatternFill(start_color="F2F2F2", fill_type="solid"),
    "MENTIONED / TO ADD": PatternFill(start_color="FFEB9C", fill_type="solid"),
    "NOT RELEVANT / OPTIONAL": PatternFill(start_color="F2F2F2", fill_type="solid"),
}

BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="0E1116", fill_type="solid")


def add_banner(ws, text, n_cols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    ws.cell(row=1, column=1, value=text).font = Font(bold=True, color="FFFFFF", size=12)
    ws.cell(row=1, column=1).fill = HEADER_FILL
    ws.cell(row=1, column=1).alignment = Alignment(wrap_text=True, vertical='center')
    ws.row_dimensions[1].height = 80


def add_header_row(ws, headers, row=2):
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical='center')
        cell.border = BORDER


def write_data_row(ws, row_idx, values, fills=None):
    for col, val in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col, value=val)
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        cell.border = BORDER
        if fills and col in fills:
            cell.fill = fills[col]


def set_widths(ws, widths):
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w


def build():
    wb = Workbook()

    # Sheet 1: Abstract Review
    ws1 = wb.active
    ws1.title = "Abstract Review"
    add_banner(ws1, BANNER_TEXT, 4)
    add_header_row(ws1, ["Field", "Status", "Review", "Rewrite / Action"], row=2)
    set_widths(ws1, [22, 20, 70, 80])
    for i, (field, status, review, rewrite) in enumerate(FIELDS, start=3):
        write_data_row(ws1, i, [field, status, review, rewrite],
                       fills={2: STATUS_FILLS.get(status, PatternFill())})

    # Sheet 2: Checklist
    ws2 = wb.create_sheet("Checklist Pre-Submit")
    add_banner(ws2, f"{PROJECT_NAME} Pre-Submission Checklist v{VERSION}", 6)
    add_header_row(ws2, ["Priority", "Item", "Field", "Effort", "Why it matters", "Status"], row=2)
    set_widths(ws2, [14, 50, 14, 12, 60, 14])
    for i, (prio, item, field, effort, why, status) in enumerate(CHECKLIST, start=3):
        write_data_row(ws2, i, [prio, item, field, effort, why, status],
                       fills={1: PRIORITY_FILLS.get(prio, PatternFill())})

    # Sheet 3: Comparables Corpus
    ws3 = wb.create_sheet("Comparables Corpus")
    add_banner(ws3, f"{PROJECT_NAME} Corpus Comparables (15 past submissions vs current)", 7)
    add_header_row(ws3, ["Project", "Round", "Budget", "Verdict", "Category", "Profile snapshot", "Lesson for current dossier"], row=2)
    set_widths(ws3, [26, 12, 14, 22, 26, 40, 50])
    for i, (proj, rnd, budget, verdict, cat, profile, lesson) in enumerate(COMPARABLES, start=3):
        write_data_row(ws3, i, [proj, rnd, budget, verdict, cat, profile, lesson],
                       fills={4: VERDICT_FILLS.get(verdict, PatternFill())})

    # Sheet 4: Integration List Audit
    ws4 = wb.create_sheet("Audit Integration List")
    add_banner(ws4, f"{PROJECT_NAME} SCF Integration List Audit", 5)
    add_header_row(ws4, ["Category", "Item", "Status in dossier", "Recommendation", "Effort"], row=2)
    set_widths(ws4, [20, 30, 22, 60, 14])
    for i, (cat, item, status, reco, effort) in enumerate(INTEGRATION_AUDIT, start=3):
        write_data_row(ws4, i, [cat, item, status, reco, effort],
                       fills={3: STATUS_AUDIT_FILLS.get(status, PatternFill())})

    wb.save(OUT)
    print(f"Saved: {OUT}")

    # QA check: count em-dashes
    from openpyxl import load_workbook
    wb2 = load_workbook(OUT)
    em = 0
    for ws in wb2.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and ", " in cell.value:
                    em += 1
    print(f"em-dash cells: {em} (expect 0)")
    print(f"Sheets: {wb2.sheetnames}")


if __name__ == "__main__":
    build()
